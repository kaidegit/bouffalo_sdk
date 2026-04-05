#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shlex
import subprocess
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import serial
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"
ANSI_ESCAPE_RE = re.compile(r"\x1b\[[0-9;?]*[ -/]*[@-~]")
GPIO_OK_PREFIX = "GPIOV OK"
GPIO_FAIL_PREFIX = "GPIOV FAIL"
DUT_USAGE_MARKER = "Usage: hbn <hbn_mode> <test_io> <trig_type> <pull>"
DUT_READY_MARKER = "GPIO Ready"
DUT_WAKEUP_RE = re.compile(r"gpio_(\d+) wakeup hbn")

PASS_FILL = PatternFill(fill_type="solid", fgColor="00B050")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
PASS_FONT = Font(bold=True, color="FFFFFF")
FAIL_FONT = Font(bold=True, color="9C0006")
HEADER_FONT = Font(bold=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

TRIGGER_PRESETS: dict[str, dict[str, Any]] = {
    "sync_fall": {
        "name": "sync_fall",
        "trig_type": 0,
        "description": "sync falling edge trigger",
        "pull": 1,
        "pull_text": "up",
        "idle_level": 1,
        "active_level": 0,
        "kind": "edge",
    },
    "sync_rise": {
        "name": "sync_rise",
        "trig_type": 1,
        "description": "sync rising edge trigger",
        "pull": 2,
        "pull_text": "down",
        "idle_level": 0,
        "active_level": 1,
        "kind": "edge",
    },
    "sync_low": {
        "name": "sync_low",
        "trig_type": 2,
        "description": "sync low level trigger",
        "pull": 1,
        "pull_text": "up",
        "idle_level": 1,
        "active_level": 0,
        "kind": "level",
    },
    "sync_high": {
        "name": "sync_high",
        "trig_type": 3,
        "description": "sync high level trigger",
        "pull": 2,
        "pull_text": "down",
        "idle_level": 0,
        "active_level": 1,
        "kind": "level",
    },
    "sync_both": {
        "name": "sync_both",
        "trig_type": 4,
        "description": "sync rising & falling edge trigger",
        "pull": 2,
        "pull_text": "down",
        "idle_level": 0,
        "active_level": 1,
        "kind": "edge",
    },
    "async_fall": {
        "name": "async_fall",
        "trig_type": 8,
        "description": "async falling edge trigger",
        "pull": 1,
        "pull_text": "up",
        "idle_level": 1,
        "active_level": 0,
        "kind": "edge",
    },
    "async_rise": {
        "name": "async_rise",
        "trig_type": 9,
        "description": "async rising edge trigger",
        "pull": 2,
        "pull_text": "down",
        "idle_level": 0,
        "active_level": 1,
        "kind": "edge",
    },
    "async_low": {
        "name": "async_low",
        "trig_type": 10,
        "description": "async low level trigger",
        "pull": 1,
        "pull_text": "up",
        "idle_level": 1,
        "active_level": 0,
        "kind": "level",
    },
    "async_high": {
        "name": "async_high",
        "trig_type": 11,
        "description": "async high level trigger",
        "pull": 2,
        "pull_text": "down",
        "idle_level": 0,
        "active_level": 1,
        "kind": "level",
    },
}


class ValidationRuntimeError(RuntimeError):
    pass


@dataclass(frozen=True)
class SerialConfig:
    label: str
    port: str
    baudrate: int
    run_reset_on_open: bool
    reset_pulse_ms: int
    boot_wait_ms: int
    dtr_low_is_true: bool
    rts_low_is_true: bool


@dataclass(frozen=True)
class BuildConfig:
    enabled: bool
    chip: str
    board: str
    extra_args: list[str]


@dataclass(frozen=True)
class FlashConfig:
    enabled: bool
    port: str
    extra_args: list[str]


@dataclass(frozen=True)
class TriggerConfig:
    name: str
    trig_type: int
    description: str
    pull: int
    pull_text: str
    idle_level: int
    active_level: int
    kind: str
    active_hold_ms: int | None


@dataclass(frozen=True)
class PinPair:
    dut_gpio: int
    dft_gpio: int


@dataclass(frozen=True)
class TimingConfig:
    startup_drain_ms: int
    dft_settle_ms: int
    ready_delay_ms: int
    edge_hold_ms: int
    level_hold_ms: int
    wakeup_timeout_ms: int
    dut_ready_timeout_ms: int
    dut_command_timeout_ms: int
    dft_command_timeout_ms: int
    post_wakeup_settle_ms: int
    quiet_after_parse_ms: int


@dataclass(frozen=True)
class RetryConfig:
    dut_command_max_retries: int
    dft_command_max_retries: int


@dataclass
class GpioCommandResult:
    ok: bool
    line: str
    elapsed_ms: int


@dataclass
class DutCommandResult:
    lines: list[str]
    ready_seen: bool
    elapsed_ms: int


@dataclass
class WakeupResult:
    passed: bool
    expected_gpio: int
    lines: list[str]
    wake_gpios: list[int]
    elapsed_ms: int
    reason: str


@dataclass
class TestRecord:
    hbn_mode: int
    dut_gpio: int
    dft_gpio: int
    trigger_name: str
    trig_type: int
    result: str
    detail: str
    dut_attempts: int
    dft_attempts: int


class ArtifactLogger:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._fp = self.path.open("w", encoding="utf-8")
        self._lock = threading.Lock()

    def log(self, channel: str, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        with self._lock:
            self._fp.write(f"[{timestamp}] [{channel}] {message}\n")
            self._fp.flush()

    def close(self) -> None:
        with self._lock:
            self._fp.close()


class SerialEndpoint:
    def __init__(self, config: SerialConfig, logger: ArtifactLogger, startup_drain_sec: float) -> None:
        self.config = config
        self.logger = logger
        self.startup_drain_sec = startup_drain_sec
        self._serial: serial.Serial | None = None
        self._lock = threading.Lock()
        self._buffer = ""

    def __enter__(self) -> "SerialEndpoint":
        self.open()
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def open(self) -> None:
        if self._serial is not None:
            return

        self.logger.log(self.config.label, f"OPEN port={self.config.port} baudrate={self.config.baudrate}")
        self._serial = serial.Serial(
            port=self.config.port,
            baudrate=self.config.baudrate,
            timeout=0.1,
            write_timeout=1,
            exclusive=True,
        )
        if self.config.run_reset_on_open:
            self.reset_to_run_mode()
        self.drain(self.startup_drain_sec)

    def close(self) -> None:
        if self._serial is None:
            return

        self.logger.log(self.config.label, "CLOSE")
        self._serial.close()
        self._serial = None

    def reset_to_run_mode(self) -> None:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        self.logger.log(
            self.config.label,
            "RESET run_mode"
            + f" dtr_low_is_true={int(self.config.dtr_low_is_true)}"
            + f" rts_low_is_true={int(self.config.rts_low_is_true)}"
            + f" reset_pulse_ms={self.config.reset_pulse_ms}"
            + f" boot_wait_ms={self.config.boot_wait_ms}",
        )
        self._set_control_lines(dtr_low=True, rts_low=True)
        time.sleep(self.config.reset_pulse_ms / 1000.0)
        self._set_control_lines(dtr_low=True, rts_low=False)
        time.sleep(self.config.boot_wait_ms / 1000.0)
        self._serial.reset_input_buffer()
        self._serial.reset_output_buffer()
        self._buffer = ""

    def drain(self, duration_sec: float) -> list[str]:
        if self._serial is None or duration_sec <= 0:
            return []

        captured: list[str] = []
        deadline = time.monotonic() + duration_sec
        while True:
            line = self._read_next_line(deadline)
            if line is None:
                break
            text = line.strip()
            if text:
                clean_text = strip_ansi(text)
                captured.append(clean_text)
                self.logger.log(self.config.label, f"RX {clean_text}")
        return captured

    def ping_gpio_tester(self, timeout_sec: float) -> GpioCommandResult:
        return self.send_gpio_command("gpiov_ping", timeout_sec)

    def send_gpio_command(self, command: str, timeout_sec: float) -> GpioCommandResult:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        with self._lock:
            self.drain(0.05)
            start = time.monotonic()
            self._write_command(command)
            deadline = start + timeout_sec
            while True:
                line = self._read_next_line(deadline)
                if line is None:
                    break

                clean_text = strip_ansi(line.strip())
                if not clean_text:
                    continue
                self.logger.log(self.config.label, f"RX {clean_text}")
                if clean_text.startswith(GPIO_OK_PREFIX) or clean_text.startswith(GPIO_FAIL_PREFIX):
                    return GpioCommandResult(
                        ok=clean_text.startswith(GPIO_OK_PREFIX),
                        line=clean_text,
                        elapsed_ms=int((time.monotonic() - start) * 1000),
                    )

        raise ValidationRuntimeError(
            f"{self.config.label} command timeout after {timeout_sec:.2f}s: {command}"
        )

    def send_dut_command(
        self,
        command: str,
        timeout_sec: float,
        required_patterns: list[str],
        success_pattern: str | None,
        failure_patterns: list[str],
        quiet_after_sec: float,
    ) -> DutCommandResult:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        with self._lock:
            self.drain(0.05)
            start = time.monotonic()
            last_rx = start
            lines: list[str] = []
            seen_required = {pattern: False for pattern in required_patterns}
            ready_seen = False
            self._write_command(command)
            deadline = start + timeout_sec

            while time.monotonic() < deadline:
                read_deadline = min(deadline, time.monotonic() + 0.05)
                line = self._read_next_line(read_deadline)
                if line is None:
                    if all(seen_required.values()) and (time.monotonic() - last_rx) >= quiet_after_sec:
                        return DutCommandResult(
                            lines=lines,
                            ready_seen=ready_seen,
                            elapsed_ms=int((time.monotonic() - start) * 1000),
                        )
                    continue

                clean_text = strip_ansi(line.strip())
                if not clean_text:
                    continue

                last_rx = time.monotonic()
                lines.append(clean_text)
                self.logger.log(self.config.label, f"RX {clean_text}")

                for pattern in required_patterns:
                    if pattern in clean_text:
                        seen_required[pattern] = True
                if success_pattern and success_pattern in clean_text:
                    ready_seen = True
                for pattern in failure_patterns:
                    if pattern in clean_text:
                        raise ValidationRuntimeError(
                            f"{self.config.label} command failed: {command} matched '{pattern}'"
                        )
                if ready_seen and all(seen_required.values()):
                    return DutCommandResult(
                        lines=lines,
                        ready_seen=True,
                        elapsed_ms=int((time.monotonic() - start) * 1000),
                    )

        missing = [pattern for pattern, seen in seen_required.items() if not seen]
        raise ValidationRuntimeError(
            f"{self.config.label} command timeout after {timeout_sec:.2f}s: {command}; missing={missing}"
        )

    def wait_for_wakeup(self, expected_gpio: int, timeout_sec: float) -> WakeupResult:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        with self._lock:
            start = time.monotonic()
            deadline = start + timeout_sec
            lines: list[str] = []
            wake_gpios: list[int] = []

            while time.monotonic() < deadline:
                line = self._read_next_line(min(deadline, time.monotonic() + 0.05))
                if line is None:
                    continue

                clean_text = strip_ansi(line.strip())
                if not clean_text:
                    continue

                lines.append(clean_text)
                self.logger.log(self.config.label, f"RX {clean_text}")

                match = DUT_WAKEUP_RE.search(clean_text)
                if match:
                    gpio = int(match.group(1))
                    wake_gpios.append(gpio)
                    if gpio == expected_gpio:
                        return WakeupResult(
                            passed=True,
                            expected_gpio=expected_gpio,
                            lines=lines,
                            wake_gpios=wake_gpios,
                            elapsed_ms=int((time.monotonic() - start) * 1000),
                            reason=f"wakeup_by_gpio_{gpio}",
                        )

        reason = "wakeup_timeout"
        if wake_gpios:
            reason = f"unexpected_wakeup_gpios={','.join(str(gpio) for gpio in wake_gpios)}"
        return WakeupResult(
            passed=False,
            expected_gpio=expected_gpio,
            lines=lines,
            wake_gpios=wake_gpios,
            elapsed_ms=int((time.monotonic() - start) * 1000),
            reason=reason,
        )

    def _write_command(self, command: str) -> None:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        payload = f"{command}\r\n".encode("utf-8")
        self.logger.log(self.config.label, f"TX {command}")
        self._serial.write(payload)
        self._serial.flush()

    def _read_next_line(self, deadline: float) -> str | None:
        if self._serial is None:
            return None

        while time.monotonic() < deadline:
            if "\n" in self._buffer:
                line, self._buffer = self._buffer.split("\n", 1)
                return line.rstrip("\r")

            waiting = self._serial.in_waiting
            chunk = self._serial.read(waiting or 1)
            if not chunk:
                continue

            self._buffer += chunk.decode("utf-8", errors="replace")

        return None

    def _set_control_lines(self, *, dtr_low: bool, rts_low: bool) -> None:
        if self._serial is None:
            return

        self._serial.dtr = dtr_low if self.config.dtr_low_is_true else not dtr_low
        self._serial.rts = rts_low if self.config.rts_low_is_true else not rts_low


def strip_ansi(text: str) -> str:
    return ANSI_ESCAPE_RE.sub("", text)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="One-click BL616CL HBN GPIO wakeup validation: DUT build + flash + DFT/DUT serial test"
    )
    parser.add_argument(
        "--config",
        default="hbn_wakeup_validation_config.example.json",
        help="Config JSON path. Default: hbn_wakeup_validation_config.example.json",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory for logs, report and result JSON. Default: validation_runs/<timestamp>",
    )
    parser.add_argument("--skip-build", action="store_true", help="Skip DUT build")
    parser.add_argument("--skip-flash", action="store_true", help="Skip DUT flash")
    parser.add_argument("--skip-test", action="store_true", help="Skip serial validation")
    return parser.parse_args()


def default_output_dir(script_dir: Path) -> Path:
    return script_dir / "validation_runs" / datetime.now().strftime(TIMESTAMP_FORMAT)


def load_json(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ValidationRuntimeError(f"Config file not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ValidationRuntimeError(f"Invalid JSON in {path}: {exc}") from exc


def parse_serial_config(section: dict[str, Any], label: str) -> SerialConfig:
    return SerialConfig(
        label=label,
        port=str(section["port"]),
        baudrate=int(section.get("baudrate", 2000000)),
        run_reset_on_open=bool(section.get("run_reset_on_open", True)),
        reset_pulse_ms=int(section.get("reset_pulse_ms", 5)),
        boot_wait_ms=int(section.get("boot_wait_ms", 200)),
        dtr_low_is_true=bool(section.get("dtr_low_is_true", True)),
        rts_low_is_true=bool(section.get("rts_low_is_true", True)),
    )


def parse_build_config(section: dict[str, Any]) -> BuildConfig:
    return BuildConfig(
        enabled=bool(section.get("enabled", True)),
        chip=str(section.get("chip", "bl616cl")),
        board=str(section.get("board", "bl616cldk")),
        extra_args=[str(item) for item in section.get("extra_args", [])],
    )


def parse_flash_config(section: dict[str, Any], default_port: str) -> FlashConfig:
    return FlashConfig(
        enabled=bool(section.get("enabled", True)),
        port=str(section.get("port", default_port)),
        extra_args=[str(item) for item in section.get("extra_args", [])],
    )


def parse_timing_config(section: dict[str, Any]) -> TimingConfig:
    return TimingConfig(
        startup_drain_ms=int(section.get("startup_drain_ms", 1500)),
        dft_settle_ms=int(section.get("dft_settle_ms", 5)),
        ready_delay_ms=int(section.get("ready_delay_ms", 10)),
        edge_hold_ms=int(section.get("edge_hold_ms", 5)),
        level_hold_ms=int(section.get("level_hold_ms", 5)),
        wakeup_timeout_ms=int(section.get("wakeup_timeout_ms", 3000)),
        dut_ready_timeout_ms=int(section.get("dut_ready_timeout_ms", 1500)),
        dut_command_timeout_ms=int(section.get("dut_command_timeout_ms", 1500)),
        dft_command_timeout_ms=int(section.get("dft_command_timeout_ms", 1000)),
        post_wakeup_settle_ms=int(section.get("post_wakeup_settle_ms", 150)),
        quiet_after_parse_ms=int(section.get("quiet_after_parse_ms", 50)),
    )


def parse_retry_config(section: dict[str, Any]) -> RetryConfig:
    return RetryConfig(
        dut_command_max_retries=int(section.get("dut_command_max_retries", 2)),
        dft_command_max_retries=int(section.get("dft_command_max_retries", 2)),
    )


def parse_trigger(raw: Any, timing: TimingConfig) -> TriggerConfig:
    if isinstance(raw, str):
        raw = {"name": raw}
    if not isinstance(raw, dict):
        raise ValidationRuntimeError(f"Unsupported trigger config: {raw!r}")

    name = raw.get("name")
    preset: dict[str, Any] = {}
    if name is not None:
        if name not in TRIGGER_PRESETS:
            raise ValidationRuntimeError(f"Unsupported trigger name: {name}")
        preset = TRIGGER_PRESETS[name].copy()
    elif "trig_type" in raw:
        trig_type = int(raw["trig_type"])
        for candidate in TRIGGER_PRESETS.values():
            if int(candidate["trig_type"]) == trig_type:
                preset = candidate.copy()
                name = str(candidate["name"])
                break
        if not preset:
            raise ValidationRuntimeError(f"Unsupported trig_type: {trig_type}")
    else:
        raise ValidationRuntimeError(f"Trigger config requires 'name' or 'trig_type': {raw}")

    merged = preset
    merged.update(raw)
    kind = str(merged.get("kind", preset["kind"]))
    active_hold_ms = merged.get("active_hold_ms")
    if active_hold_ms is None:
        active_hold_ms = timing.level_hold_ms if kind == "level" else timing.edge_hold_ms

    return TriggerConfig(
        name=str(merged["name"]),
        trig_type=int(merged["trig_type"]),
        description=str(merged["description"]),
        pull=int(merged["pull"]),
        pull_text=str(merged["pull_text"]),
        idle_level=int(merged["idle_level"]),
        active_level=int(merged["active_level"]),
        kind=kind,
        active_hold_ms=int(active_hold_ms),
    )


def parse_pins(raw_pins: list[Any]) -> list[PinPair]:
    pairs: list[PinPair] = []
    for item in raw_pins:
        if isinstance(item, dict):
            pairs.append(PinPair(dut_gpio=int(item["dut_gpio"]), dft_gpio=int(item["dft_gpio"])))
            continue
        raise ValidationRuntimeError(f"Unsupported pin mapping: {item!r}")
    return pairs


def log_progress(logger: ArtifactLogger, stage: str, message: str) -> None:
    logger.log(stage, message)
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] [{stage}] {message}")


def run_subprocess(cmd: list[str], cwd: Path, logger: ArtifactLogger, channel: str) -> None:
    logger.log(channel, f"RUN cwd={cwd} cmd={shlex.join(cmd)}")
    proc = subprocess.Popen(
        cmd,
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    assert proc.stdout is not None
    for line in proc.stdout:
        text = line.rstrip()
        if text:
            logger.log(channel, text)
            print(f"[{channel}] {text}")
    return_code = proc.wait()
    if return_code != 0:
        raise ValidationRuntimeError(f"{channel} failed with exit code {return_code}: {shlex.join(cmd)}")


def build_dut(project_dir: Path, build: BuildConfig, logger: ArtifactLogger) -> None:
    cmd = ["make", f"CHIP={build.chip}", f"BOARD={build.board}", *build.extra_args]
    log_progress(logger, "build", f"start chip={build.chip} board={build.board}")
    run_subprocess(cmd, project_dir, logger, "build")
    log_progress(logger, "build", "done")


def flash_dut(project_dir: Path, build: BuildConfig, flash: FlashConfig, logger: ArtifactLogger) -> None:
    cmd = ["make", "flash", f"CHIP={build.chip}", f"COMX={flash.port}", *flash.extra_args]
    log_progress(logger, "flash", f"start chip={build.chip} port={flash.port}")
    run_subprocess(cmd, project_dir, logger, "flash")
    log_progress(logger, "flash", "done")


def send_gpio_with_retry(
    endpoint: SerialEndpoint,
    command: str,
    timeout_sec: float,
    retries: int,
    logger: ArtifactLogger,
) -> tuple[GpioCommandResult, int]:
    last_error: Exception | None = None
    for attempt in range(1, retries + 2):
        try:
            result = endpoint.send_gpio_command(command, timeout_sec)
            if not result.ok:
                raise ValidationRuntimeError(result.line)
            return result, attempt
        except Exception as exc:
            last_error = exc
            logger.log(endpoint.config.label, f"RETRY_GPIO command={command} attempt={attempt} error={exc}")
            time.sleep(0.1)
    raise ValidationRuntimeError(f"{endpoint.config.label} command failed after retries: {command}; {last_error}")


def ensure_dut_ready(
    endpoint: SerialEndpoint,
    timing: TimingConfig,
    retries: int,
    logger: ArtifactLogger,
) -> None:
    last_error: Exception | None = None
    for attempt in range(1, retries + 2):
        try:
            endpoint.send_dut_command(
                command="hbn",
                timeout_sec=timing.dut_ready_timeout_ms / 1000.0,
                required_patterns=[DUT_USAGE_MARKER],
                success_pattern=DUT_USAGE_MARKER,
                failure_patterns=["this event has no process"],
                quiet_after_sec=timing.quiet_after_parse_ms / 1000.0,
            )
            return
        except Exception as exc:
            last_error = exc
            logger.log(endpoint.config.label, f"RETRY_DUT_READY attempt={attempt} error={exc}")
            endpoint.reset_to_run_mode()
            endpoint.drain(max(endpoint.startup_drain_sec, timing.post_wakeup_settle_ms / 1000.0))
    raise ValidationRuntimeError(f"DUT not ready on {endpoint.config.port}: {last_error}")


def arm_dut_hbn(
    endpoint: SerialEndpoint,
    hbn_mode: int,
    pin: int,
    trigger: TriggerConfig,
    timing: TimingConfig,
    retries: int,
    logger: ArtifactLogger,
) -> tuple[DutCommandResult, int]:
    command = f"hbn {hbn_mode} {pin} {trigger.trig_type} {trigger.pull}"
    required = [
        f"hbn_mode: HBN_{hbn_mode}",
        f"test_io: {pin}",
        f"test_trig_type:{trigger.description}",
        f"pull: {trigger.pull_text}",
    ]
    failure_patterns = ["hbn_mode must be less than", "test_io must be less than", "pull must be less than", "[ERR]", DUT_USAGE_MARKER]
    last_error: Exception | None = None
    for attempt in range(1, retries + 2):
        try:
            result = endpoint.send_dut_command(
                command=command,
                timeout_sec=timing.dut_command_timeout_ms / 1000.0,
                required_patterns=required,
                success_pattern=DUT_READY_MARKER,
                failure_patterns=failure_patterns,
                quiet_after_sec=timing.quiet_after_parse_ms / 1000.0,
            )
            return result, attempt
        except Exception as exc:
            last_error = exc
            logger.log(endpoint.config.label, f"RETRY_DUT_CMD command={command} attempt={attempt} error={exc}")
            endpoint.reset_to_run_mode()
            endpoint.drain(max(endpoint.startup_drain_sec, timing.post_wakeup_settle_ms / 1000.0))
            ensure_dut_ready(endpoint, timing, 0, logger)
    raise ValidationRuntimeError(f"DUT command failed after retries: {command}; {last_error}")


def run_single_case(
    dut: SerialEndpoint,
    dft: SerialEndpoint,
    hbn_mode: int,
    pair: PinPair,
    trigger: TriggerConfig,
    timing: TimingConfig,
    retries: RetryConfig,
    logger: ArtifactLogger,
    case_index: int,
    total_cases: int,
) -> TestRecord:
    start_message = (
        f"case={case_index}/{total_cases} hbn={hbn_mode} dut_gpio={pair.dut_gpio} "
        f"dft_gpio={pair.dft_gpio} trigger={trigger.name} trig_type={trigger.trig_type} "
        f"idle={trigger.idle_level} active={trigger.active_level} hold_ms={trigger.active_hold_ms}"
    )
    log_progress(logger, "test", f"{start_message} start")

    dft_attempts = 0
    dut_attempts = 0

    try:
        ensure_dut_ready(dut, timing, retries.dut_command_max_retries, logger)

        _, attempt = send_gpio_with_retry(
            dft,
            f"gpiov_config {pair.dft_gpio} out init={trigger.idle_level}",
            timing.dft_command_timeout_ms / 1000.0,
            retries.dft_command_max_retries,
            logger,
        )
        dft_attempts = max(dft_attempts, attempt)
        time.sleep(timing.dft_settle_ms / 1000.0)

        _, attempt = send_gpio_with_retry(
            dft,
            f"gpiov_write {pair.dft_gpio} {trigger.idle_level}",
            timing.dft_command_timeout_ms / 1000.0,
            retries.dft_command_max_retries,
            logger,
        )
        dft_attempts = max(dft_attempts, attempt)
        time.sleep(timing.dft_settle_ms / 1000.0)

        dut_result, attempt = arm_dut_hbn(
            dut,
            hbn_mode,
            pair.dut_gpio,
            trigger,
            timing,
            retries.dut_command_max_retries,
            logger,
        )
        dut_attempts = max(dut_attempts, attempt)
        time.sleep(timing.ready_delay_ms / 1000.0)

        _, attempt = send_gpio_with_retry(
            dft,
            f"gpiov_write {pair.dft_gpio} {trigger.active_level}",
            timing.dft_command_timeout_ms / 1000.0,
            retries.dft_command_max_retries,
            logger,
        )
        dft_attempts = max(dft_attempts, attempt)
        time.sleep(trigger.active_hold_ms / 1000.0)

        _, attempt = send_gpio_with_retry(
            dft,
            f"gpiov_write {pair.dft_gpio} {trigger.idle_level}",
            timing.dft_command_timeout_ms / 1000.0,
            retries.dft_command_max_retries,
            logger,
        )
        dft_attempts = max(dft_attempts, attempt)

        wakeup = dut.wait_for_wakeup(pair.dut_gpio, timing.wakeup_timeout_ms / 1000.0)
        time.sleep(timing.post_wakeup_settle_ms / 1000.0)

        detail = (
            f"{wakeup.reason}; dut_ready={int(dut_result.ready_seen)}; "
            f"dut_lines={len(dut_result.lines)}; wake_lines={len(wakeup.lines)}"
        )
        result = "PASS" if wakeup.passed else "FAIL"
        log_progress(logger, "test", f"{start_message} result={result} detail={detail}")
        return TestRecord(
            hbn_mode=hbn_mode,
            dut_gpio=pair.dut_gpio,
            dft_gpio=pair.dft_gpio,
            trigger_name=trigger.name,
            trig_type=trigger.trig_type,
            result=result,
            detail=detail,
            dut_attempts=dut_attempts,
            dft_attempts=dft_attempts,
        )
    except Exception as exc:
        detail = str(exc)
        log_progress(logger, "test", f"{start_message} result=FAIL detail={detail}")
        try:
            dut.reset_to_run_mode()
            dut.drain(max(dut.startup_drain_sec, timing.post_wakeup_settle_ms / 1000.0))
        except Exception as reset_exc:
            logger.log("test", f"DUT_RECOVERY_FAIL error={reset_exc}")
            detail = f"{detail}; dut_recovery={reset_exc}"
        return TestRecord(
            hbn_mode=hbn_mode,
            dut_gpio=pair.dut_gpio,
            dft_gpio=pair.dft_gpio,
            trigger_name=trigger.name,
            trig_type=trigger.trig_type,
            result="FAIL",
            detail=detail,
            dut_attempts=dut_attempts,
            dft_attempts=dft_attempts,
        )


def run_validation(
    dut: SerialEndpoint,
    dft: SerialEndpoint,
    hbn_modes: list[int],
    pins: list[PinPair],
    triggers: list[TriggerConfig],
    timing: TimingConfig,
    retries: RetryConfig,
    logger: ArtifactLogger,
) -> list[TestRecord]:
    records: list[TestRecord] = []
    total_cases = len(hbn_modes) * len(pins) * len(triggers)
    case_index = 0

    ping_result, _ = send_gpio_with_retry(
        dft,
        "gpiov_ping",
        timing.dft_command_timeout_ms / 1000.0,
        retries.dft_command_max_retries,
        logger,
    )
    logger.log("DFT", f"PING_OK elapsed_ms={ping_result.elapsed_ms}")
    ensure_dut_ready(dut, timing, retries.dut_command_max_retries, logger)

    for hbn_mode in hbn_modes:
        for pair in pins:
            for trigger in triggers:
                case_index += 1
                records.append(
                    run_single_case(
                        dut=dut,
                        dft=dft,
                        hbn_mode=hbn_mode,
                        pair=pair,
                        trigger=trigger,
                        timing=timing,
                        retries=retries,
                        logger=logger,
                        case_index=case_index,
                        total_cases=total_cases,
                    )
                )
    return records


def build_result_workbook(records: list[TestRecord], output_path: Path, trigger_order: list[str]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    if not records:
        summary_sheet["A1"] = "No validation records generated"
        summary_sheet["A1"].alignment = CELL_ALIGNMENT
        summary_sheet.column_dimensions["A"].width = 32
        workbook.save(output_path)
        return

    workbook.remove(summary_sheet)

    order_map = {name: index for index, name in enumerate(trigger_order)}
    grouped: dict[int, list[TestRecord]] = {}
    for record in records:
        grouped.setdefault(record.hbn_mode, []).append(record)

    for hbn_mode in sorted(grouped):
        sheet = workbook.create_sheet(f"hbn{hbn_mode}")
        headers = ["GPIO", "Wakeup Mode", "Result", "Detail", "DFT GPIO", "Trig Type", "DUT Attempts", "DFT Attempts"]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CELL_ALIGNMENT

        sheet.freeze_panes = "A2"
        row = 2
        rows = sorted(grouped[hbn_mode], key=lambda item: (item.dut_gpio, order_map[item.trigger_name]))
        by_gpio: dict[int, list[TestRecord]] = {}
        for record in rows:
            by_gpio.setdefault(record.dut_gpio, []).append(record)

        for gpio in sorted(by_gpio):
            items = by_gpio[gpio]
            start_row = row
            for record in items:
                sheet.cell(row=row, column=1, value=record.dut_gpio)
                sheet.cell(row=row, column=2, value=record.trigger_name)
                result_cell = sheet.cell(row=row, column=3, value=record.result)
                sheet.cell(row=row, column=4, value=record.detail)
                sheet.cell(row=row, column=5, value=record.dft_gpio)
                sheet.cell(row=row, column=6, value=record.trig_type)
                sheet.cell(row=row, column=7, value=record.dut_attempts)
                sheet.cell(row=row, column=8, value=record.dft_attempts)

                for column in range(1, 9):
                    sheet.cell(row=row, column=column).alignment = CELL_ALIGNMENT

                if record.result == "PASS":
                    result_cell.fill = PASS_FILL
                    result_cell.font = PASS_FONT
                else:
                    result_cell.fill = FAIL_FILL
                    result_cell.font = FAIL_FONT
                row += 1

            end_row = row - 1
            if end_row > start_row:
                sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        widths = [12, 18, 12, 60, 12, 12, 14, 14]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(output_path)


def write_results_json(records: list[TestRecord], output_path: Path) -> None:
    payload = {
        "summary": {
            "total": len(records),
            "pass": sum(1 for record in records if record.result == "PASS"),
            "fail": sum(1 for record in records if record.result != "PASS"),
        },
        "records": [
            {
                "hbn_mode": record.hbn_mode,
                "dut_gpio": record.dut_gpio,
                "dft_gpio": record.dft_gpio,
                "trigger_name": record.trigger_name,
                "trig_type": record.trig_type,
                "result": record.result,
                "detail": record.detail,
                "dut_attempts": record.dut_attempts,
                "dft_attempts": record.dft_attempts,
            }
            for record in records
        ],
    }
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = script_dir / config_path

    raw = load_json(config_path)
    build = parse_build_config(raw.get("build", {}))
    serial_section = raw.get("serial", {})
    dut_serial = parse_serial_config(serial_section["dut"], "DUT")
    dft_serial = parse_serial_config(serial_section["dft"], "DFT")
    flash = parse_flash_config(raw.get("flash", {}), dut_serial.port)
    timing = parse_timing_config(raw.get("timing", {}))
    retries = parse_retry_config(raw.get("retries", {}))
    tests = raw.get("tests", {})
    hbn_modes = [int(item) for item in tests.get("hbn_modes", [0, 1])]
    pins = parse_pins(list(tests.get("pins", [])))
    triggers = [parse_trigger(item, timing) for item in tests.get("triggers", list(TRIGGER_PRESETS.keys()))]
    if not pins:
        raise ValidationRuntimeError("Config must define at least one pin mapping under tests.pins")

    output_dir = Path(args.output_dir) if args.output_dir else default_output_dir(script_dir)
    if not output_dir.is_absolute():
        output_dir = script_dir / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    log_path = output_dir / "hbn_wakeup_validation.log"
    report_path = output_dir / "hbn_wakeup_validation_report.xlsx"
    result_json_path = output_dir / "hbn_wakeup_validation_results.json"
    logger = ArtifactLogger(log_path)

    try:
        log_progress(logger, "config", f"config={config_path}")
        log_progress(logger, "config", f"output_dir={output_dir}")

        if build.enabled and not args.skip_build:
            build_dut(script_dir, build, logger)
        else:
            log_progress(logger, "build", "skipped")

        if flash.enabled and not args.skip_flash:
            flash_dut(script_dir, build, flash, logger)
        else:
            log_progress(logger, "flash", "skipped")

        records: list[TestRecord] = []
        if not args.skip_test:
            with SerialEndpoint(dut_serial, logger, timing.startup_drain_ms / 1000.0) as dut, SerialEndpoint(
                dft_serial, logger, timing.startup_drain_ms / 1000.0
            ) as dft:
                records = run_validation(
                    dut=dut,
                    dft=dft,
                    hbn_modes=hbn_modes,
                    pins=pins,
                    triggers=triggers,
                    timing=timing,
                    retries=retries,
                    logger=logger,
                )
        else:
            log_progress(logger, "test", "skipped")

        write_results_json(records, result_json_path)
        build_result_workbook(records, report_path, [trigger.name for trigger in triggers])

        passed = sum(1 for record in records if record.result == "PASS")
        failed = len(records) - passed
        log_progress(logger, "summary", f"pass={passed} fail={failed} total={len(records)}")
        log_progress(logger, "summary", f"report={report_path}")
        log_progress(logger, "summary", f"json={result_json_path}")
        return 0
    finally:
        logger.close()


if __name__ == "__main__":
    raise SystemExit(main())
